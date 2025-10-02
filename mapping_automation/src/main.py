# --- Imports ---
import psycopg2.pool
from Automation_Scripts import db_creds

import requests
import json
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Example of generic field mapping definitions
field_mapping_definitions = {
    "IS_ACTIVE": {
        "long_name": "StatusFlag",
        "transformation": "IF(StatusFlag=''Active'',1,0)"
    }
}


pool = None  # global placeholder

def create_pool():
    return psycopg2.pool.SimpleConnectionPool(
        minconn=1,
        maxconn=10,
        database=db_creds.DB_MAIN,
        host=db_creds.DB_HOST,
        user=db_creds.DB_USER,
        password=db_creds.DB_PASS,
        port=db_creds.DB_PORT
    )

def get_connection():
    global pool
    if pool is None:
        pool = create_pool()
    return pool.getconn()


# --- Base Data Collection ---
def get_src_info(cursor, src_list, dl_type):
    srcs_str = "', '".join(src_list)
    qry = f"""  select  info.source, info.protocol, info.provider, cls.dataset_id, cls.dataset_name, cls.dataset_description, '{dl_type}' AS download_type
                from table_dataset_config cls
                        join table_source_info info on info.id = cls.dataset_id
                where info.source in ('{srcs_str}')
                        and cls.download_type = '{dl_type}'
                ;"""

    cursor.execute(qry)
    return  [item for item in cursor.fetchall()]


def get_field_info(cursor, fields, dl_type):
    field_str = "', '".join(fields)
    qry = f"""  select id, name
                from table_canonical_fields
                where download_type = '{dl_type}'
                        and name in ('{field_str}');"""

    cursor.execute(qry)
    return [item for item in cursor.fetchall()]


# --- Mapping Audit & Excel Write ---
def mapping_audit(cursor, tup_list):
    updated_list = []

    for i in tup_list:
        field_id = i[7]
        dataset_id = i[3]
        dataset_name = i[4]
        download_type = i[6]

        qry = f"""  select is_active
                    from table_mapping
                    where field_id = '{field_id}'
                            and dataset_id = '{dataset_id}'
                            and dataset_name = '{dataset_name}'
                            and download_type = '{download_type}';"""

        cursor.execute(qry)
        result = cursor.fetchall()

        if not result:
            updated_list.append(i + ('Not Mapped',))
        else:
            active_status = result[0][0]
            if active_status:
                updated_list.append(i + ('Mapped',))
            else:
                updated_list.append(i + ('Deactivated',))

    return updated_list


def append_proposed_fields(audit_data, field_mapping_definitions):
    updated_data = []
    for row in audit_data:
        status = row[9]
        canonical_field = row[8]
        if status != 'Mapped' and canonical_field in field_mapping_definitions:
            long_name = field_mapping_definitions[canonical_field].get("long_name", "")
            transformation = field_mapping_definitions[canonical_field].get("transformation", "")
        else:
            long_name = "N/A"
            transformation = "N/A"
        updated_data.append(row + (long_name, transformation))
    return updated_data


def get_metadata_elastic_search(source, dataset_name, field_name, resource, auth_url):
    query = {
        "_source": ["documentId", "className", "longName", "tableSystemName"],
        "query": {
            "bool": {
                "must": [
                    {"term": {"documentId": {"value": source.lower()}}},
                    {"term": {"className": {"value": dataset_name.lower()}}},
                    {"match_phrase": {"longName": field_name}}
                ]
            }
        }
    }

    if resource:
        query["query"]["bool"]["must"].insert(2, {"term": {"resource": {"value": resource.lower()}}})

    headers = {"Content-Type": "application/json"}
    try:
        response = requests.get(auth_url, headers=headers, data=json.dumps(query))
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        return {"error": str(e)}


def elasticsearch_check_from_df(df, auth_url):
    updated_rows = []

    for _, row in df.iterrows():
        status = row['Mapping Status']
        if status == 'Mapped':
            row['es_Pass'] = 'N/A'
            row['Proposed Fields Long Name'] = 'N/A'
            updated_rows.append(row)
            continue

        source = row['Source']
        dataset_name = row['Class']
        protocol = row['Protocol']
        download_type = row['Download Type']
        proposed_fields = row['Proposed Field Short Name']

        if not proposed_fields or pd.isna(proposed_fields):
            row['es_Pass'] = 'N'
            row['Proposed Fields Long Name'] = 'NF'
            updated_rows.append(row)
            continue

        download_type_lower = download_type.lower()
        if 'listing' in download_type_lower:
            resource = "Property" if protocol == "RETS" else "EntityType" if protocol == "WEBAPI" else ""
        elif download_type_lower == "openhouse":
            resource = "OpenHouse" if protocol == "RETS" else "EntityType" if protocol == "WEBAPI" else ""
        elif download_type_lower in {"agent", "office"}:
            resource = None
        else:
            resource = ""

        fields = [f.strip() for f in str(proposed_fields).split(',')]
        all_found = True
        long_names = []

        for field in fields:
            metadata = get_metadata_elastic_search(source, dataset_name, field, resource, auth_url)
            hits = metadata.get("hits", {}).get("hits", [])
            if hits:
                table_name = hits[0]['_source'].get('tableSystemName', 'NF')
                long_names.append(table_name)
            else:
                long_names.append('NF')
                all_found = False

        row['es_Pass'] = 'Y' if all_found else 'N'
        row['Proposed Fields Long Name'] = ','.join(long_names)
        updated_rows.append(row)

    return pd.DataFrame(updated_rows)


def add_finalized_transformation(df):
    finalized = []
    for _, row in df.iterrows():
        if row.get('es_Pass') == 'Y':
            short_names = [s.strip() for s in str(row.get('Proposed Field Short Name', '')).split(',')]
            long_names = [l.strip() for l in str(row.get('Proposed Fields Long Name', '')).split(',')]
            mapping = dict(zip(short_names, long_names))
            transformation = row.get('Proposed Transformation', '')
            for short, long in mapping.items():
                transformation = transformation.replace(short, long)
            row['Finalized Transformation'] = transformation
        else:
            row['Finalized Transformation'] = 'N/A'
        finalized.append(row)
    return pd.DataFrame(finalized)


def write_updated_audit_to_excel(headers, rows, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Results"

    ws.append(headers)
    for row in rows:
        ws.append(row)

    for col_idx, col in enumerate(headers, 1):
        max_length = max(len(str(cell)) for cell in [col] + [r[col_idx - 1] for r in rows])
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    table_ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    tab = Table(displayName="AuditTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(file_path)
    print(f"Excel file '{file_path}' has been created successfully.")


# --- Insert Statement Generators ---
def canonical_inserts_from_df(df, conn, download_type):
    inserts = []
    cursor = conn.cursor()

    for _, row in df.iterrows():
        field_id = row['Field ID']
        dataset_id = row['Dataset ID']
        dataset_name = row['Class']
        dataset_desc = row['Class Description']
        mapping = row['Finalized Transformation']

        check_qry = f"""
        SELECT 1 FROM table_mapping
        WHERE field_id = {field_id}
        AND dataset_id = {dataset_id}
        AND dataset_name = '{dataset_name}'
        AND download_type = '{download_type}';
        """
        cursor.execute(check_qry)
        if cursor.fetchone():
            print(f"Skipping existing mapping: field_id={field_id}, dataset_id={dataset_id}, dataset_name={dataset_name}")
            continue

        insert_stmt = f"""
        INSERT INTO table_mapping
        (field_id, dataset_id, column_transformation_id, custom_transformation, is_active, last_update_ts, create_ts, download_type, dataset_name, dataset_description, auto_mapped)
        VALUES ({field_id}, {dataset_id}, 3, '{mapping}', true, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, '{download_type}', '{dataset_name}', '{dataset_desc}', true);
        """
        inserts.append(insert_stmt)

    for stmt in inserts:
        cursor.execute(stmt)
    conn.commit()

    if inserts:
        print("Canonical Inserts Created")
    else:
        print("No new canonical inserts created")


def origin_inserts_from_df(df, conn):
    inserts = []
    cursor = conn.cursor()

    for _, row in df.iterrows():
        field_id = row['Field ID']
        dataset_id = row['Dataset ID']
        dataset_name = row['Class']
        short_names = [s.strip() for s in str(row['Proposed Fields Long Name']).split(',')]
        long_names = [l.strip() for l in str(row['Proposed Field Short Name']).split(',')]

        mapping_id_qry = f"""
            SELECT id, dataset_name FROM table_mapping
            WHERE field_id = {field_id} AND dataset_id = {dataset_id};
        """
        cursor.execute(mapping_id_qry)
        results = cursor.fetchall()

        if not results:
            print(f"No mapping IDs found for field_id={field_id}, dataset_id={dataset_id}")
            continue

        matched = False
        for mapping_id, db_class in results:
            if db_class != dataset_name:
                continue

            matched = True
            for short_name, long_name in zip(short_names, long_names):
                check_qry = f"""
                    SELECT 1 FROM table_origin_field
                    WHERE mapping_id = {mapping_id}
                    AND source_field = '{short_name}'
                    AND dataset_id = {dataset_id};
                """
                cursor.execute(check_qry)
                if cursor.fetchone():
                    print(f"Skipping existing origin field: mapping_id={mapping_id}, source_field={short_name}, dataset_id={dataset_id}")
                    continue

                origin_stmt = f"""
                    INSERT INTO table_origin_field
                    (mapping_id, source_field, dataset_id, is_active, last_update_ts, create_ts, short_name, long_name)
                    VALUES ({mapping_id}, '{short_name}', {dataset_id}, true, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, '{short_name}', '{long_name}');
                """
                inserts.append(origin_stmt)

        if not matched:
            print(f"No matching dataset found for field_id={field_id}, dataset_id={dataset_id}, dataset_name={dataset_name}")

    for stmt in inserts:
        cursor.execute(stmt)
    conn.commit()

    if inserts:
        print("Origin Inserts Created")
    else:
        print("No new origin inserts created")


# --- Update Statement Generators ---
def canonical_updates_from_df(df, conn):
    cursor = conn.cursor()

    updates_executed = False

    for _, row in df.iterrows():
        field_id = row['Field ID']
        dataset_id = row['Dataset ID']
        dataset_name = row['Class']
        download_type = row['Download Type']
        new_transformation = row['Finalized Transformation']

        qry = f"""
        SELECT custom_transformation FROM table_mapping
        WHERE field_id = {field_id}
        AND dataset_id = {dataset_id}
        AND dataset_name = '{dataset_name}'
        AND download_type = '{download_type}';
        """
        cursor.execute(qry)
        result = cursor.fetchone()

        if result:
            current_transformation = result[0]
            if current_transformation.strip() == new_transformation.strip():
                update_stmt = f"""
                UPDATE table_mapping
                SET is_active = true, last_update_ts = CURRENT_TIMESTAMP
                WHERE field_id = {field_id}
                AND dataset_id = {dataset_id}
                AND dataset_name = '{dataset_name}'
                AND download_type = '{download_type}';
                """
            else:
                update_stmt = f"""
                UPDATE table_mapping
                SET custom_transformation = '{new_transformation}',
                    is_active = true,
                    last_update_ts = CURRENT_TIMESTAMP
                WHERE field_id = {field_id}
                AND dataset_id = {dataset_id}
                AND dataset_name = '{dataset_name}'
                AND download_type = '{download_type}';
                """
            cursor.execute(update_stmt)
            updates_executed = True
        else:
            print("Mapping not found")

    conn.commit()

    if updates_executed:
        print("Canonical Updates Created")
    else:
        print("No updates executed")


def origin_updates_from_df(df, conn):
    cursor = conn.cursor()
    updates_executed = False

    for _, row in df.iterrows():
        field_id = row['Field ID']
        dataset_id = row['Dataset ID']
        short_names = [s.strip() for s in str(row['Proposed Fields Short Name']).split(',')]
        long_names = [l.strip() for l in str(row['Proposed Fields Long Name']).split(',')]

        mapping_id_qry = f"""
        SELECT id FROM table_mapping
        WHERE field_id = {field_id} AND dataset_id = {dataset_id};
        """
        cursor.execute(mapping_id_qry)
        result = cursor.fetchone()
        if not result:
            print(f"Mapping ID not found for field_id={field_id}, dataset_id={dataset_id}")
            continue
        mapping_id = result[0]

        for short_name, long_name in zip(short_names, long_names):
            check_qry = f"""
            SELECT 1 FROM table_origin_field
            WHERE mapping_id = {mapping_id}
            AND source_field = '{short_name}'
            AND dataset_id = {dataset_id};
            """
            cursor.execute(check_qry)
            if cursor.fetchone():
                # Row exists, update it
                update_stmt = f"""
                UPDATE table_origin_field
                SET is_active = true, last_update_ts = CURRENT_TIMESTAMP
                WHERE mapping_id = {mapping_id}
                AND source_field = '{short_name}'
                AND dataset_id = {dataset_id};
                """
            else:
                # Row does not exist, insert it
                update_stmt = f"""
                INSERT INTO table_origin_field
                (mapping_id, source_field, dataset_id, is_active, last_update_ts, create_ts, short_name, long_name)
                VALUES ({mapping_id}, '{short_name}', {dataset_id}, true, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, '{short_name}', '{long_name}');
                """
            cursor.execute(update_stmt)
            updates_executed = True

    conn.commit()

    if updates_executed:
        print("Origin Updates Created")
    else:
        print("No updates executed")


# --- Main Execution ---
def main():
    source_list = ['SRC_A', 'SRC_B', 'SRC_C']
    download_type = 'agent'
    canonical_fields = tuple(field_mapping_definitions.keys())
    auth_url = "https://placeholder-opensearch-url.com/api/search"
    out_path = '/path/to/output/'
    out_file_name = f"{out_path}Canonical_Audit_{download_type}_results.xlsx"

    conn = get_connection()
    cursor = conn.cursor()

    source_info = get_src_info(cursor, source_list, download_type)
    field_info = get_field_info(cursor, canonical_fields, download_type)

    master_list = [l1 + l2 for l1 in source_info for l2 in field_info]
    audit_tups = mapping_audit(cursor, master_list)
    audit_tups_with_proposals = append_proposed_fields(audit_tups, field_mapping_definitions)

    initial_headers = ['Source', 'Protocol', 'Provider', 'Dataset ID', 'Class', 'Class Description', 'Download Type',
                       'Field ID', 'Canonical Field Name', 'Mapping Status', 'Proposed Fields Short Name',
                       'Proposed Transformation']
    audit_df = pd.DataFrame(audit_tups_with_proposals, columns=initial_headers)

    # Run Elasticsearch check and add 'es_Pass' and 'Proposed Fields Long Name'
    audit_df_with_es = elasticsearch_check_from_df(audit_df, auth_url)

    audit_df_with_es = add_finalized_transformation(audit_df_with_es)

    # Final headers for Excel output
    final_headers = initial_headers + ['es_Pass', 'Proposed Fields Long Name', 'Finalized Transformation']

    # Write final audit to Excel
    write_updated_audit_to_excel(final_headers, audit_df_with_es.values.tolist(), out_file_name)

    # Pause and prompt user to review the spreadsheet
    input(
        f"\n✅ Audit spreadsheet saved to '{out_file_name}'. Please review before continuing.\nPress Enter to proceed...")

    # Generate Inserts and Updates for 'Not Mapped' Records
    unmapped_df = audit_df_with_es[
        (audit_df_with_es['Mapping Status'] == 'Not Mapped') &
        (audit_df_with_es['es_Pass'] == 'Y')
        ]
    if not unmapped_df.empty:
        print("\n--- Canonical Insert Statements ---")
        canonical_inserts_from_df(unmapped_df, conn, download_type)

        print("\n--- Origin Insert Statements ---")
        origin_inserts_from_df(unmapped_df, conn)

    # Generate Updates for 'Deactivated' Records with valid metadata
    deactivated_df = audit_df_with_es[
        (audit_df_with_es['Mapping Status'] == 'Deactivated') &
        (audit_df_with_es['es_Pass'] == 'Y')
        ]
    if not deactivated_df.empty:
        print("\n--- Canonical Update Statements ---")
        canonical_updates_from_df(deactivated_df, conn)

        print("\n--- Origin Update Statements ---")
        origin_updates_from_df(deactivated_df, conn)

    cursor.close()
    conn.close()


if __name__ == "__main__":
    main()
